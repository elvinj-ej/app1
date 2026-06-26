"""
aws_ingestor.py
---------------
Parses an AWS CUR file (pivoted Excel format) and upserts rows into cur_data.

Expected input format (same as AWSCUR tab):
  Col A: account_id
  Col B: account_name
  Col C: workloads_tag
  Col D: outcomegroup_tag
  Col E: category          ('monthly_expense' | 'marketplace')
  Col F onwards: month columns (datetime values, one per calendar month)

Summary rows (account_id is blank/None) are skipped automatically.
\t prefixes on account IDs are stripped.

Returns a dict: { rows_upserted, months_found, errors }
"""

import os
import re
import logging
from datetime import datetime
from aws_db import get_conn

log = logging.getLogger(__name__)

VALID_CATEGORIES = {"monthly_expense", "marketplace"}


def _parse_excel(path: str) -> tuple[list[datetime], list[dict]]:
    """
    Parse the pivoted CUR Excel file.
    Returns (month_dates, data_rows) where data_rows is a list of dicts:
      { account_id, account_name, workloads_tag, outcomegroup, category,
        amounts: {month_str: float} }
    """
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    month_dates: list[datetime] = []
    month_col_indices: list[int] = []
    data_rows: list[dict] = []

    for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
        if row_idx == 0:
            # Header row — identify date columns (index 5 onwards)
            for col_idx, val in enumerate(row):
                if col_idx < 5:
                    continue
                if isinstance(val, datetime):
                    month_dates.append(val)
                    month_col_indices.append(col_idx)
            continue

        # Skip summary rows (account_id is blank)
        raw_account_id = row[0]
        if raw_account_id is None or str(raw_account_id).strip() in ("", "None"):
            continue

        # Strip \t artifacts from copy-paste
        account_id = re.sub(r"[\t\s]+", "", str(raw_account_id)).strip()
        if not account_id:
            continue

        account_name  = (row[1] or "").strip()
        workloads_tag = (row[2] or "").strip() or None
        outcomegroup  = (row[3] or "").strip() or None
        category      = (row[4] or "").strip().lower()

        if category not in VALID_CATEGORIES:
            continue

        amounts: dict[str, float] = {}
        for col_idx, dt in zip(month_col_indices, month_dates):
            val = row[col_idx] if col_idx < len(row) else None
            if val is not None:
                try:
                    amounts[dt.strftime("%Y-%m-01")] = float(val)
                except (TypeError, ValueError):
                    pass

        if not amounts:
            continue

        data_rows.append({
            "account_id":    account_id,
            "account_name":  account_name,
            "workloads_tag": workloads_tag,
            "outcomegroup":  outcomegroup,
            "category":      category,
            "amounts":       amounts,
        })

    wb.close()
    return month_dates, data_rows


def _parse_csv(path: str) -> tuple[list[datetime], list[dict]]:
    """
    Same logic but for a plain CSV CUR export.
    First 5 columns: account_id, account_name, workloads_tag, outcomegroup_tag, category
    Remaining columns: date headers in 'YYYY-MM-DD' or 'YYYY-MM' format.
    """
    import csv

    month_dates: list[datetime] = []
    month_col_indices: list[int] = []
    data_rows: list[dict] = []

    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        for row_idx, row in enumerate(reader):
            if row_idx == 0:
                for col_idx, val in enumerate(row):
                    if col_idx < 5:
                        continue
                    val = val.strip()
                    for fmt in ("%Y-%m-%d", "%Y-%m", "%Y/%m/%d", "%Y/%m"):
                        try:
                            dt = datetime.strptime(val, fmt)
                            month_dates.append(dt)
                            month_col_indices.append(col_idx)
                            break
                        except ValueError:
                            pass
                continue

            raw_account_id = row[0].strip() if row else ""
            if not raw_account_id:
                continue
            account_id = re.sub(r"[\t\s]+", "", raw_account_id).strip()
            if not account_id:
                continue

            account_name  = row[1].strip() if len(row) > 1 else ""
            workloads_tag = row[2].strip() if len(row) > 2 else None
            outcomegroup  = row[3].strip() if len(row) > 3 else None
            category      = row[4].strip().lower() if len(row) > 4 else ""

            if category not in VALID_CATEGORIES:
                continue

            amounts: dict[str, float] = {}
            for col_idx, dt in zip(month_col_indices, month_dates):
                val = row[col_idx].strip() if col_idx < len(row) else ""
                if val:
                    try:
                        amounts[dt.strftime("%Y-%m-01")] = float(val.replace(",", ""))
                    except ValueError:
                        pass

            if amounts:
                data_rows.append({
                    "account_id":    account_id,
                    "account_name":  account_name,
                    "workloads_tag": workloads_tag or None,
                    "outcomegroup":  outcomegroup or None,
                    "category":      category,
                    "amounts":       amounts,
                })

    return month_dates, data_rows


def ingest_cur(path: str, uploaded_by: str = "") -> dict:
    """
    Parse the CUR file at `path` and upsert all rows into the DB.
    Returns { rows_upserted, months_found, errors }
    """
    ext = os.path.splitext(path)[1].lower()
    errors: list[str] = []

    try:
        if ext in (".xlsx", ".xlsm", ".xls"):
            month_dates, data_rows = _parse_excel(path)
        elif ext == ".csv":
            month_dates, data_rows = _parse_csv(path)
        else:
            return {"rows_upserted": 0, "months_found": 0,
                    "errors": [f"Unsupported file type: {ext}"]}
    except Exception as e:
        log.error(f"CUR parse error: {e}")
        return {"rows_upserted": 0, "months_found": 0, "errors": [str(e)]}

    if not data_rows:
        return {"rows_upserted": 0, "months_found": len(month_dates),
                "errors": ["No data rows found — check file format."]}

    rows_upserted = 0
    with get_conn() as conn:
        for row in data_rows:
            for month_str, amount in row["amounts"].items():
                conn.execute("""
                    INSERT INTO cur_data
                        (account_id, account_name, workloads_tag, outcomegroup, category, month, amount)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                    ON CONFLICT(account_id, COALESCE(workloads_tag,''), category, month)
                    DO UPDATE SET amount=excluded.amount,
                                  account_name=excluded.account_name,
                                  outcomegroup=excluded.outcomegroup
                """, (
                    row["account_id"], row["account_name"],
                    row["workloads_tag"], row["outcomegroup"],
                    row["category"], month_str, amount,
                ))
                rows_upserted += 1

        conn.execute("""
            INSERT INTO upload_log (filename, rows_upserted, months_found, uploaded_by)
            VALUES (?, ?, ?, ?)
        """, (os.path.basename(path), rows_upserted, len(month_dates), uploaded_by))

    log.info(f"CUR ingest complete: {rows_upserted} rows, {len(month_dates)} months, file={path}")
    return {
        "rows_upserted": rows_upserted,
        "months_found":  len(month_dates),
        "months":        [dt.strftime("%Y-%m-01") for dt in month_dates],
        "errors":        errors,
    }
