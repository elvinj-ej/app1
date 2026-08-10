"""
Parse AWS CUR file (CSV or Excel) and upsert into cur_data.

CSV/Excel format (pivoted):
  Col A: account_id   (may have leading \\t prefix in exports)
  Col B: account_name
  Col C: workloads_tag
  Col D: outcomegroup_tag
  Col E: category     ('monthly_expense' | 'marketplace')
  Col F+: month columns

Summary rows (blank account_id) are skipped.
Re-uploading the same period is safe — rows are upserted with snapshot backup.
"""
import os
import random
import re
import logging
from datetime import datetime
from aws_db import get_conn

log = logging.getLogger(__name__)

VALID_CATEGORIES = {"monthly_expense", "marketplace"}

# Workload categories to prefer for validation cross-check
_CHECK_COST_CATEGORIES = ("Project",)


def _clean_account_id(raw):
    s = str(raw).strip()
    s = s.lstrip("\\t").strip()
    return re.sub(r"[\t\s]+", "", s)


def _parse_date(val):
    if isinstance(val, datetime):
        return val
    s = str(val).strip()
    for fmt in ("%Y-%m-%d", "%Y-%m", "%Y/%m/%d", "%Y/%m", "%b-%Y", "%B-%Y", "%b %Y"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            pass
    return None


def _parse_excel(path):
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    month_dates, month_cols, data_rows = [], [], []

    for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
        if row_idx == 0:
            for col_idx, val in enumerate(row):
                if col_idx >= 5:
                    dt = _parse_date(val) if val is not None else None
                    if dt:
                        month_dates.append(dt)
                        month_cols.append(col_idx)
            continue

        if row[0] is None or str(row[0]).strip() in ("", "None"):
            continue
        account_id = _clean_account_id(row[0])
        if not account_id:
            continue

        category = (row[4] or "").strip().lower()
        if category not in VALID_CATEGORIES:
            continue

        amounts = {}
        for col_idx, dt in zip(month_cols, month_dates):
            val = row[col_idx] if col_idx < len(row) else None
            if val is not None:
                try:
                    amounts[dt.strftime("%Y-%m-01")] = float(val)
                except (TypeError, ValueError):
                    pass

        if amounts:
            data_rows.append({
                "account_id":    account_id,
                "account_name":  (row[1] or "").strip(),
                "workloads_tag": (row[2] or "").strip(),
                "outcomegroup":  (row[3] or "").strip() or None,
                "category":      category,
                "amounts":       amounts,
            })

    wb.close()
    return month_dates, data_rows


def _parse_csv(path):
    import csv

    month_dates, month_cols, data_rows = [], [], []

    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        for row_idx, row in enumerate(reader):
            if row_idx == 0:
                for col_idx, val in enumerate(row):
                    if col_idx >= 5:
                        dt = _parse_date(val)
                        if dt:
                            month_dates.append(dt)
                            month_cols.append(col_idx)
                continue

            raw_id = row[0].strip() if row else ""
            account_id = _clean_account_id(raw_id)
            if not account_id:
                continue

            category = row[4].strip().lower() if len(row) > 4 else ""
            if category not in VALID_CATEGORIES:
                continue

            amounts = {}
            for col_idx, dt in zip(month_cols, month_dates):
                val = row[col_idx].strip() if col_idx < len(row) else ""
                if val:
                    try:
                        amounts[dt.strftime("%Y-%m-01")] = float(val.replace(",", ""))
                    except ValueError:
                        pass

            if amounts:
                data_rows.append({
                    "account_id":    account_id,
                    "account_name":  row[1].strip() if len(row) > 1 else "",
                    "workloads_tag": row[2].strip() if len(row) > 2 else "",
                    "outcomegroup":  row[3].strip() if len(row) > 3 else None,
                    "category":      category,
                    "amounts":       amounts,
                })

    return month_dates, data_rows


def parse_cur_only(path):
    """Parse a CUR file and return (month_dates, data_rows) without touching the DB."""
    ext = os.path.splitext(path)[1].lower()
    if ext in (".xlsx", ".xlsm", ".xls"):
        return _parse_excel(path)
    elif ext == ".csv":
        return _parse_csv(path)
    else:
        raise ValueError(f"Unsupported file type: {ext}")


def validate_cur_data(data_rows, month_dates):
    """
    Sanity-check historical months in the new file against what's already in the DB.

    Picks a random Project/ADA workload and compares up to 5 prior months
    (excluding the latest month in the file). Returns a dict with:
      ok            — bool: True if all checked months match
      workload      — which workload was checked
      months_checked — list of months compared
      matches       — list of {month, file_total, db_total}
      mismatches    — list of {month, file_total, db_total, diff}
      skipped       — bool: True if validation was skipped (not enough history)
      skip_reason   — why it was skipped (if skipped)
    """
    if not month_dates:
        return {"ok": True, "skipped": True, "skip_reason": "No months in file"}

    all_months = sorted({dt.strftime("%Y-%m-01") for dt in month_dates})
    if len(all_months) < 2:
        return {"ok": True, "skipped": True,
                "skip_reason": "Only one month in file — nothing to validate against"}

    latest_month   = all_months[-1]
    history_months = all_months[:-1]
    check_months   = history_months[-5:]  # up to 5 prior months

    # Aggregate file data: (workloads_tag, month) → total across all accounts & categories
    file_agg = {}
    for row in data_rows:
        wl = row["workloads_tag"]
        for m, amt in row["amounts"].items():
            if m in check_months:
                key = (wl, m)
                file_agg[key] = file_agg.get(key, 0.0) + amt

    # Prefer Project/ADA workloads from the DB
    with get_conn() as conn:
        preferred = {
            r["name"] for r in conn.execute(
                "SELECT name FROM workloads WHERE cost_category IN ('Project')"
            ).fetchall()
        }
        # Fallback: any workload present in the DB
        all_db_wls = {
            r["workloads_tag"] for r in conn.execute(
                "SELECT DISTINCT workloads_tag FROM cur_data WHERE month IN ({})".format(
                    ",".join("?" * len(check_months))
                ), check_months
            ).fetchall()
        }

    file_workloads = {wl for (wl, _) in file_agg}
    candidates = list(preferred & file_workloads & all_db_wls)
    if not candidates:
        candidates = list(file_workloads & all_db_wls)
    if not candidates:
        return {"ok": True, "skipped": True,
                "skip_reason": "No overlapping workloads found between file and existing DB data"}

    checked_wl = random.choice(sorted(candidates))

    # Fetch DB totals for checked workload over check months
    with get_conn() as conn:
        db_rows = conn.execute(
            "SELECT month, SUM(amount) as total FROM cur_data "
            "WHERE workloads_tag=? AND month IN ({}) GROUP BY month".format(
                ",".join("?" * len(check_months))
            ),
            [checked_wl] + list(check_months)
        ).fetchall()
    db_totals = {r["month"]: r["total"] for r in db_rows}

    matches, mismatches = [], []
    for m in check_months:
        file_val = file_agg.get((checked_wl, m))
        db_val   = db_totals.get(m)
        if file_val is None or db_val is None:
            continue
        diff = abs(file_val - db_val)
        entry = {"month": m, "file_total": round(file_val, 2), "db_total": round(db_val, 2)}
        if diff > 0.05:
            mismatches.append({**entry, "diff": round(diff, 2)})
        else:
            matches.append(entry)

    if not matches and not mismatches:
        return {"ok": True, "skipped": True,
                "skip_reason": f"Workload '{checked_wl}' has no overlapping months with DB"}

    return {
        "ok":             len(mismatches) == 0,
        "skipped":        False,
        "workload":       checked_wl,
        "latest_month":   latest_month,
        "months_checked": check_months,
        "matches":        matches,
        "mismatches":     mismatches,
    }


def preview_cur_changes(data_rows, month_dates):
    """
    For each month in the file, compute the total amounts (all workloads combined)
    and compare against what's already in the DB. Returns a list of:
      {month, file_total, db_total, delta, is_new}
    sorted by month, so the user can see what will change.
    """
    if not month_dates:
        return []

    all_months = sorted({dt.strftime("%Y-%m-01") for dt in month_dates})

    # Aggregate file totals per month
    file_month_totals = {}
    for row in data_rows:
        for m, amt in row["amounts"].items():
            file_month_totals[m] = file_month_totals.get(m, 0.0) + amt

    # Fetch DB totals per month
    with get_conn() as conn:
        db_rows = conn.execute(
            "SELECT month, SUM(amount) as total FROM cur_data "
            "WHERE month IN ({}) GROUP BY month".format(
                ",".join("?" * len(all_months))
            ), all_months
        ).fetchall()
    db_month_totals = {r["month"]: r["total"] for r in db_rows}

    result = []
    for m in all_months:
        f = file_month_totals.get(m, 0.0)
        d = db_month_totals.get(m, None)
        result.append({
            "month":      m,
            "file_total": round(f, 2),
            "db_total":   round(d, 2) if d is not None else None,
            "delta":      round(f - d, 2) if d is not None else None,
            "is_new":     d is None,
        })
    return result


def ingest_cur(path, uploaded_by=""):
    ext = os.path.splitext(path)[1].lower()
    try:
        month_dates, data_rows = parse_cur_only(path)
    except Exception as e:
        log.error(f"CUR parse error: {e}")
        return {"rows_upserted": 0, "months_found": 0, "errors": [str(e)]}

    if not data_rows:
        return {
            "rows_upserted": 0,
            "months_found": len(month_dates),
            "errors": ["No data rows found — check file format. Expected columns: account_id, account_name, workloads_tag, outcomegroup_tag, category, then month columns."],
        }

    rows_upserted = 0
    with get_conn() as conn:
        # Insert upload_log first so we have an ID for snapshots
        conn.execute(
            "INSERT INTO upload_log (filename, rows_upserted, months_found, uploaded_by) VALUES (?,?,?,?)",
            (os.path.basename(path), 0, len(month_dates), uploaded_by),
        )
        upload_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]

        # Collect all (account_id, workloads_tag, category, month) tuples from incoming data
        incoming_keys = []
        for row in data_rows:
            for month_str in row["amounts"]:
                incoming_keys.append((row["account_id"], row["workloads_tag"], row["category"], month_str))

        # Fetch existing rows for those keys in one query (batch by chunks if large)
        existing = {}
        chunk_size = 400  # stay well under SQLite's 999-variable limit (4 vars per row)
        for i in range(0, len(incoming_keys), chunk_size):
            chunk = incoming_keys[i:i + chunk_size]
            placeholders = ",".join("(?,?,?,?)" for _ in chunk)
            flat = [v for key in chunk for v in key]
            rows = conn.execute(
                "SELECT account_id, workloads_tag, category, month, amount, account_name, outcomegroup "
                "FROM cur_data WHERE (account_id, workloads_tag, category, month) IN (" + placeholders + ")",
                flat
            ).fetchall()
            for r in rows:
                existing[(r["account_id"], r["workloads_tag"], r["category"], r["month"])] = r

        # Write snapshots
        for row in data_rows:
            for month_str in row["amounts"]:
                key = (row["account_id"], row["workloads_tag"], row["category"], month_str)
                old = existing.get(key)
                conn.execute(
                    "INSERT OR IGNORE INTO cur_data_snapshots "
                    "(upload_id, account_id, workloads_tag, category, month, account_name, outcomegroup, old_amount) "
                    "VALUES (?,?,?,?,?,?,?,?)",
                    (upload_id, row["account_id"], row["workloads_tag"], row["category"], month_str,
                     old["account_name"] if old else None,
                     old["outcomegroup"] if old else None,
                     old["amount"] if old else None),  # None = was a new row
                )

        # Upsert data
        for row in data_rows:
            for month_str, amount in row["amounts"].items():
                conn.execute(
                    "INSERT INTO cur_data "
                    "(account_id, account_name, workloads_tag, outcomegroup, category, month, amount) "
                    "VALUES (?,?,?,?,?,?,?) "
                    "ON CONFLICT(account_id, workloads_tag, category, month) "
                    "DO UPDATE SET amount=excluded.amount, account_name=excluded.account_name, outcomegroup=excluded.outcomegroup",
                    (row["account_id"], row["account_name"], row["workloads_tag"],
                     row["outcomegroup"], row["category"], month_str, amount),
                )
                rows_upserted += 1

        # Update upload_log with final row count
        conn.execute("UPDATE upload_log SET rows_upserted=? WHERE id=?", (rows_upserted, upload_id))

    log.info(f"CUR ingest: {rows_upserted} rows, {len(month_dates)} months — {path} (upload_id={upload_id})")
    return {
        "upload_id":     upload_id,
        "rows_upserted": rows_upserted,
        "months_found":  len(month_dates),
        "months":        [dt.strftime("%Y-%m-01") for dt in month_dates],
        "errors":        [],
    }
